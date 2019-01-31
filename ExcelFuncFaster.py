def BugzillaExcelPop(Reference_file,Updated_File,List_Attributes=[]):
    from Test3func import bugzilla
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(Reference_file)
    updated_wb = openpyxl.load_workbook(Updated_File)
    #print(wb.get_sheet_names())
    sheet = wb.get_sheet_by_name('Sheet1')
    updated_sheet = updated_wb.active
    cell=2
    while (sheet['A'+str(cell)].value!=None):
        updated_sheet.cell(row=cell,column=1).value= sheet['A'+str(cell)].value
        bugresult=bugzilla(sheet['A' + str(cell)].value)
        for item in List_Attributes:
            if (item.lower() != "duplicates"):
                updated_sheet.cell(row=cell, column=(List_Attributes.index(item) + 2)).value =bugresult[item.lower()]
            else:
                updated_sheet.cell(row=cell, column=(List_Attributes.index(item) + 2)).value = str(bugresult[item.lower()])

        cell=cell+1
    updated_wb.save(Updated_File)

