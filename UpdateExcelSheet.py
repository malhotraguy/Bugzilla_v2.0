import openpyxl
from openpyxl.styles import Font
from ExcelFuncFaster import BugzillaExcelPop
wb=openpyxl.Workbook() #Opening the excel workbook
#wb.save('BugReports_updated.xlsx')

sheet=wb.active #bring the active to sheet "sheet"
colum=["Bug ID","Product","Component","Creation_time","Last_change_time","Duplicates"] #list of fields to be written in first row of Workbook
for colNum in range(1,7):
    sheet.cell(row=1, column=colNum).font = Font(size=11,name='Calibri', bold=True)#giving styling to each cell
    sheet.cell(row=1,column=colNum).value=colum[colNum-1]#writing value to each cell

wb.save('BugReports_updated.xlsx')#saving the workbook
BugzillaExcelPop("BugReports.xlsx",'BugReports_updated.xlsx',colum[1:6])
